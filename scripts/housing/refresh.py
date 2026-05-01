"""Refresh site/housing-map.html from the GTKSA housing survey.

Data sources (first one that works wins):
1. HOUSING_SHEET_CSV_URL env var — published-CSV URL from Google Sheets
   (File → Share → Publish to web → CSV → copy URL).
2. HOUSING_XLSX env var — path to a local .xlsx export.
3. Default Dropbox path on the maintainer's Mac (dev fallback).

Strips the contact column before serializing. Uses geo_cache.json to avoid
re-calling Nominatim for known addresses; new addresses are looked up and
appended to the cache. Apartments with no clean address fall back to a
generic "Atlanta, GA" pin so they still appear (flagged approx).

Run from anywhere: paths are resolved relative to this script.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import sys
import time
import urllib.request
import urllib.parse
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
SITE_DIR = REPO_ROOT
SCRIPT_DIR = Path(__file__).resolve().parent
CACHE_PATH = SCRIPT_DIR / 'geo_cache.json'
OUTPUT_HTML = SITE_DIR / 'housing-map.html'

DEFAULT_XLSX = Path(
    '/Users/hkim3239/GaTech Dropbox/Hyunju Kim/atlanta_homepage/'
    '[GTKSA] 2026 거주지 정보 조사 리스트.xlsx'
)

PII_FIELDS = {'문의/Refer 연락'}

# ---------------------------------------------------------------------------
# Manual overrides — edit by hand when a new apartment shows up that
# Nominatim doesn't geocode well (or when a typo'd address needs help).
# ---------------------------------------------------------------------------

HARDCODE_COORDS = {
    # canonical key -> (lat, lon, label)
    'tenth and home': (33.7818, -84.3954, '251 10th St NW, Atlanta, GA 30332'),
    'paloma west midtown': (33.7917, -84.4136, 'Paloma West Midtown, 1062 Brady Ave NW, Atlanta, GA 30318'),
    'rambler atlanta': (33.7849, -84.3849, 'Rambler Atlanta, 1080 Peachtree St NE, Atlanta, GA 30309'),
    'camden brookwood': (33.8088, -84.3856, 'Camden Brookwood, 2500 Peachtree Rd NW, Atlanta, GA 30305'),
    'maa lenox': (33.8466, -84.3621, 'MAA Lenox, 3324 Peachtree Rd NE, Atlanta, GA 30326'),
    'maa centennial park': (33.7607, -84.3934, 'MAA Centennial Park, 285 Centennial Olympic Park Dr NW, Atlanta, GA 30313'),
    'alexander at the district': (33.8845, -84.4690, 'The Alexander at the District, Cumberland, Atlanta, GA'),
    'imt buckhead on 26th': (33.8265, -84.3676, 'IMT Buckhead, 2698 Sidney Marcus Blvd NE, Atlanta, GA 30324'),
    'reserve at lenox park': (33.8635, -84.3372, 'Reserve at Lenox Park, 3535 Roswell Rd NE, Atlanta, GA 30305'),
    'steelworks': (33.7867, -84.4070, 'Steelworks, 1220 Mecaslin St NW, Atlanta, GA 30318'),
    'momentum midtown': (33.7841, -84.3897, 'Momentum Midtown, 1018 W Peachtree St NW, Atlanta, GA 30309'),
    'tens on west': (33.7837, -84.3895, 'Tens on West, 1010 W Peachtree St NW, Atlanta, GA 30309'),
    'society atlanta': (33.7793, -84.3833, 'Society Atlanta, 811 Peachtree St NE, Atlanta, GA 30308'),
    'sora at spring quarter': (33.7825, -84.3892, 'Sora at Spring Quarter, 1000 Spring St NW, Atlanta, GA 30309'),
    # Nominatim mis-hits — pinned by hand
    'alexan on 8th': (33.78131, -84.38510, 'Alexan on 8th, 880 Peachtree St NE, Atlanta, GA 30309'),
    'the baxter decatur': (33.78195, -84.29685, 'The Baxter Decatur, 1605 Church St, Decatur, GA 30033'),
    'mix': (33.75696, -84.38260, 'The Mix, 120 Piedmont Ave NE, Atlanta, GA 30303'),
    'emmi': (33.78684, -84.38962, 'EMMI, 1405 Spring St NW, Atlanta, GA 30309'),
    'brooke': (33.90934, -84.27595, 'The Brooke, 2500 Shallowford Rd, Chamblee, GA'),
    'arlo': (33.77380, -84.29611, 'Arlo, 245 E Trinity Pl, Decatur, GA 30030'),
    'gateway at cedar brook': (33.80132, -84.28482, 'Gateway at Cedar Brook, 3117 Cedar Brook Dr, Decatur, GA 30033'),
    'maa brookhaven': (33.86438, -84.33970, 'MAA Brookhaven, 2829 Caldwell Rd NE, Atlanta, GA 30319'),
    'weatherstone condo': (33.82340, -84.35020, 'Weatherstone Condo, 1266 Weatherstone Dr NE, Atlanta, GA 30324'),
    'abberly skye': (33.80553, -84.28232, 'Abberly Skye, 2550 Blackmon Dr, Decatur, GA 30033'),
    'avana on main': (33.82010, -84.35050, 'Avana on Main, 508 Main St NE, Atlanta, GA 30324'),
    'overton rise': (33.87807, -84.46692, 'Overton Rise, 3695 Cumberland Blvd SE, Atlanta, GA 30339'),
    # Vague unnamed houses — approximate pins
    '__house__:marietta clipper lane': (33.9526, -84.5499, 'Clipper Ln, Marietta, GA (approx.)'),
    '__house__:piedmont ave': (33.7800, -84.3760, 'Piedmont Ave, Atlanta, GA (approx.)'),
    '__house__:buford': (34.1209, -84.0058, 'Buford, GA (approx.)'),
    '__house__:atlantic station': (33.7916, -84.3996, 'Atlantic Station, Atlanta, GA (approx.)'),
}

NEIGHBORHOOD = {
    'hadley': 'Midtown',
    'trace midtown': 'Midtown',
    'society atlanta': 'Midtown',
    '903 peachtree': 'Midtown',
    'hub atlanta': 'Midtown',
    'sora at spring quarter': 'Midtown',
    'linea midtown': 'Midtown',
    'momentum midtown': 'Midtown',
    'tens on west': 'Midtown',
    'rambler atlanta': 'Midtown',
    'alexan on 8th': 'Midtown',
    'nine15 midtown': 'Midtown',
    'hanover midtown': 'Midtown',
    'metropolis': 'Midtown',
    'ascent midtown': 'Midtown',
    'emmi': 'Midtown',
    'windsor at midtown': 'Midtown',
    '100midtown': 'Midtown',
    '__house__:piedmont ave': 'Midtown',

    'bower westside': 'West Midtown',
    'paloma west midtown': 'West Midtown',
    'skyline west': 'West Midtown',
    'steelworks': 'West Midtown',
    'avana westside': 'West Midtown',
    'osprey': 'West Midtown',
    'brady': 'West Midtown',

    'exchange': 'Atlantic Station / Home Park',
    '__house__:atlantic station': 'Atlantic Station / Home Park',
    'local on 14th': 'Atlantic Station / Home Park',
    'westside union': 'Atlantic Station / Home Park',
    'tenth and home': 'Atlantic Station / Home Park',

    'maa centennial park': 'Downtown',
    'mix': 'Downtown',
    # everything else falls through to '그 외'
}

# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def fetch_csv(url: str) -> list[dict]:
    req = urllib.request.Request(url, headers={'User-Agent': 'gtksa-housing-refresh/1.0'})
    with urllib.request.urlopen(req, timeout=30) as resp:
        text = resp.read().decode('utf-8')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise SystemExit('CSV had no rows')
    headers = [h.strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({headers[i]: (r[i].strip() if i < len(r) else '') for i in range(len(headers))})
    return out


def fetch_xlsx(path: Path) -> list[dict]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise SystemExit('openpyxl required for xlsx fallback: pip install openpyxl')
    import datetime as dt
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [(ws.cell(row=1, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, dt.datetime):
                v = v.strftime('%Y-%m-%d')
            d[h] = '' if v is None else (str(v).strip() if not isinstance(v, str) else v.strip())
        if any(d.values()):
            out.append(d)
    return out


def deduplicate_headers(headers: list[str]) -> list[str]:
    seen, out = {}, []
    for h in headers:
        h = h or ''
        if h in seen:
            seen[h] += 1
            out.append(f'{h}__{seen[h]}')
        else:
            seen[h] = 0
            out.append(h)
    return out


def load_responses() -> tuple[list[dict], list[str]]:
    url = os.environ.get('HOUSING_SHEET_CSV_URL', '').strip()
    if url:
        print(f'[refresh] fetching CSV: {url[:80]}…')
        raw = fetch_csv(url)
    else:
        xlsx_path = os.environ.get('HOUSING_XLSX', '').strip()
        path = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX
        if not path.exists():
            raise SystemExit(
                f'No HOUSING_SHEET_CSV_URL set, and no XLSX at {path}.\n'
                'Set HOUSING_SHEET_CSV_URL to a Google-Sheets-published CSV URL, '
                'or HOUSING_XLSX to a local .xlsx path.'
            )
        print(f'[refresh] reading xlsx: {path}')
        raw = fetch_xlsx(path)

    if not raw:
        raise SystemExit('No rows in source data')

    # Normalize duplicate headers (form has two "월 수입 증빙 필요 유무" cols, etc.)
    headers = deduplicate_headers(list(raw[0].keys()))
    out = []
    for i, row in enumerate(raw, start=2):
        # Re-key with disambiguated headers based on column order
        keys_in = list(row.keys())
        d = {headers[j]: row[keys_in[j]] for j in range(len(headers))}
        if any(v not in (None, '', 'None') for v in d.values()):
            d['_row'] = i
            out.append(d)
    return out, headers


# ---------------------------------------------------------------------------
# Normalization (port from extract.py)
# ---------------------------------------------------------------------------

def norm_name(name: str) -> str:
    if not name:
        return ''
    n = name.lower().strip()
    n = re.sub(r'\s+', ' ', n)
    n = n.replace('apartments', '').replace('apartment', '').replace('&', 'and')
    n = re.sub(r'\s+(bldg|building)\b.*$', '', n)
    n = re.sub(r'^gt\s+', '', n)
    n = re.sub(r'\btenth.*home.*$', 'tenth and home', n)
    n = re.sub(r'^tenth [a-z]$', 'tenth and home', n)
    if n in ('on campus housing', 'gt on campus housing', 'glc'):
        n = 'tenth and home'
    n = re.sub(r'^tens on west.*$', 'tens on west', n)
    n = re.sub(r'\bsociety atlanta\b.*$', 'society atlanta', n)
    n = re.sub(r'\b소사이어티\b', 'society atlanta', n)
    n = re.sub(r'^the (hadley|exchange|brady|baxter decatur|brooke|local on 14th|ridge|mix)$', r'\1', n)
    n = re.sub(r'^trace.*$', 'trace midtown', n)
    n = re.sub(r'^hub.*$', 'hub atlanta', n)
    n = re.sub(r'^ridge$', 'the ridge', n)
    n = re.sub(r'\bsora.*spring quarter\b', 'sora at spring quarter', n)
    n = re.sub(r'^bower westside.*$', 'bower westside', n)
    n = re.sub(r'^paloma west midtown.*$', 'paloma west midtown', n)
    n = re.sub(r'^903 peachtree.*$', '903 peachtree', n)
    n = re.sub(r'^linea midtown.*$', 'linea midtown', n)
    n = re.sub(r'^the local on 14th.*$', 'the local on 14th', n)
    n = re.sub(r'^camden buckhead square.*$', 'camden buckhead square', n)
    n = re.sub(r'^momentum midtown.*$', 'momentum midtown', n)
    n = re.sub(r'^steelworks.*$', 'steelworks', n)
    return n.strip()


def short_type(t: str) -> str:
    if not t:
        return 'Unknown'
    if '(Student Housing)' in t:
        return 'Student Housing'
    if '아파트 (일반)' in t:
        return 'Apartment'
    if '기숙사' in t:
        return 'Dorm'
    if '주택' in t:
        return 'House'
    return 'Unknown'


# ---------------------------------------------------------------------------
# Geocoding (with cache)
# ---------------------------------------------------------------------------

def load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text())
    return {}


def save_cache(cache: dict) -> None:
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True))


ADDR_OK = re.compile(r'^\s*\d+\s+\S')


def pick_addr(addrs: list[str]) -> str | None:
    cleaned = []
    for a in addrs:
        if not a:
            continue
        s = str(a).strip()
        if s.startswith('http') or any(k in s for k in ['추가', '후가', '구글', '없음']):
            continue
        cleaned.append(s)
    cleaned.sort(key=lambda s: (0 if ADDR_OK.match(s) else 1, -len(s)))
    return cleaned[0] if cleaned else None


def nominatim(query: str) -> tuple[float, float, str] | None:
    url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode(
        {'q': query, 'format': 'json', 'limit': 1, 'countrycodes': 'us'}
    )
    req = urllib.request.Request(url, headers={'User-Agent': 'gtksa-housing-refresh/1.0 (gtksa.net)'})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
    except Exception as e:
        print(f'[geo] nominatim error: {e}')
        return None
    if not data:
        return None
    return float(data[0]['lat']), float(data[0]['lon']), data[0].get('display_name', '')


def geocode_groups(groups: dict, cache: dict) -> dict:
    """Return {key: {lat, lon, address, source}}. Updates cache in-place."""
    geo = {}
    for key, rows in groups.items():
        if key in HARDCODE_COORDS:
            lat, lon, label = HARDCODE_COORDS[key]
            geo[key] = {'lat': lat, 'lon': lon, 'address': label, 'source': 'hardcode'}
            continue
        if key in cache:
            geo[key] = cache[key]
            continue
        addrs = sorted({(r.get('거주지 주소') or '').strip() for r in rows})
        addr = pick_addr(addrs)
        if not addr:
            # Best-effort: use first non-empty address even if messy, else fallback to Atlanta center
            messy = next((a for a in addrs if a), 'Atlanta, GA')
            geo[key] = {'lat': 33.7756, 'lon': -84.3963, 'address': f'{messy} (대략적 위치)', 'source': 'fallback-noaddr'}
            cache[key] = geo[key]
            continue
        q = addr if ('GA' in addr or 'Georgia' in addr.lower()) else f'{addr}, Atlanta, GA'
        res = nominatim(q)
        time.sleep(1.1)
        if res:
            lat, lon, dn = res
            geo[key] = {'lat': lat, 'lon': lon, 'address': addr, 'source': 'nominatim'}
        else:
            print(f'[geo] miss for {key!r} ({addr!r}); using Atlanta-center fallback')
            geo[key] = {'lat': 33.7756, 'lon': -84.3963, 'address': f'{addr} (대략적 위치)', 'source': 'fallback'}
        cache[key] = geo[key]
    return geo


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def build_locations(rows: list[dict], geo: dict) -> list[dict]:
    groups: dict[str, list[dict]] = {}
    for r in rows:
        groups.setdefault(r['_norm_name'], []).append(r)

    locs = []
    for key, group_rows in groups.items():
        if key not in geo:
            continue
        g = geo[key]
        name_counts = Counter()
        for r in group_rows:
            n = (r.get('거주지(아파트) 이름') or '').strip()
            if n and n != '없음':
                name_counts[n.lower()] += 1
        if name_counts:
            winner_lower = max(name_counts, key=lambda k: (name_counts[k], len(k)))
            cands = [r['거주지(아파트) 이름'].strip() for r in group_rows
                     if (r.get('거주지(아파트) 이름') or '').strip().lower() == winner_lower]
            display = max(cands, key=len) if cands else winner_lower.title()
        else:
            display = '(이름 없음 / House)'
        locs.append({
            'key': key,
            'name': display,
            'address': g['address'],
            'lat': g['lat'],
            'lon': g['lon'],
            'approx': g.get('source') in ('hardcode', 'fallback', 'fallback-noaddr') and key.startswith('__house__:')
                      or g.get('source') in ('fallback', 'fallback-noaddr'),
            'neighborhood': NEIGHBORHOOD.get(key, '그 외'),
            'rows': [{k: v for k, v in r.items() if k not in PII_FIELDS and not k.startswith('_')} | {'_row': r['_row']}
                     for r in group_rows],
        })
    locs.sort(key=lambda d: (-len(d['rows']), d['name'].lower()))
    return locs


def render_html(locations: list[dict], headers: list[str]) -> str:
    payload = {
        'locations': locations,
        'headers': [h for h in headers if h not in PII_FIELDS],
    }
    template_path = SCRIPT_DIR / 'template.html'
    template = template_path.read_text(encoding='utf-8')
    total_rows = sum(len(loc['rows']) for loc in locations)
    return (template
        .replace('__PAYLOAD__', json.dumps(payload, ensure_ascii=False))
        .replace('__NRESP__', str(total_rows))
        .replace('__NLOC__', str(len(locations)))
    )


def main() -> int:
    rows, headers = load_responses()
    print(f'[refresh] loaded {len(rows)} rows')

    # Filter to rows that have any data
    cleaned = []
    for r in rows:
        name = (r.get('거주지(아파트) 이름') or '').strip()
        addr = (r.get('거주지 주소') or '').strip()
        nname = norm_name(name)
        if not nname or nname in ('없음', '(no name)'):
            nname = '__house__:' + (addr.lower().strip() or f'row{r["_row"]}')
        r['_norm_name'] = nname
        cleaned.append(r)

    groups: dict[str, list[dict]] = {}
    for r in cleaned:
        groups.setdefault(r['_norm_name'], []).append(r)
    print(f'[refresh] {len(groups)} unique groups')

    cache = load_cache()
    geo = geocode_groups(groups, cache)
    save_cache(cache)
    print(f'[refresh] geocoded {len(geo)} / {len(groups)}')

    locations = build_locations(cleaned, geo)
    html = render_html(locations, headers)
    OUTPUT_HTML.write_text(html, encoding='utf-8')
    print(f'[refresh] wrote {OUTPUT_HTML} ({len(html)} bytes)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
